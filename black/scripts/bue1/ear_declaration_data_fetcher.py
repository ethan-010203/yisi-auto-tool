import json
import os
import re
import sys
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from playwright.sync_api import (
    BrowserContext,
    Locator,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)


EAR_PORTAL_URL = "https://www.ear-system.de/ear-portal/"
LOGIN_FRAME_PATTERN = re.compile(r"/ear-portal/secure/common/portalaufgaben\.jsf")
IST_INPUT_PATH_FRAGMENT = "/ear-portal/secure/hersteller/ist-inputmeldung/ist-inputmeldungen.jsf"
SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm"}

HEADER_AUTHORIZED_REP = "授权代表\nbevollmächtigter Vertreter"
HEADER_WEEE_NUMBER = "WEEE号\nWEEE-Nummer"
HEADER_COMPANY_NAME_CN = "中文名\nFirmenname auf Chinesisch"
HEADER_COMPANY_NAME_EN = "英文名\nFirmenname auf Englisch"
HEADER_CATEGORY = "类别\nKategorie"
HEADER_GERMAN_CATEGORY = "德语类目"
HEADER_ACCOUNT = "账号"
HEADER_PASSWORD = "密码"
HEADER_DECLARED_WEIGHT = "3月申报数据"
HEADER_FETCHED_WEIGHT = "官网上抓取的数据（3月）"
MISSING_MATCH_WRITEBACK_VALUE = "查无此数据"

EXPECTED_HEADERS = [
    HEADER_AUTHORIZED_REP,
    HEADER_WEEE_NUMBER,
    HEADER_COMPANY_NAME_CN,
    HEADER_COMPANY_NAME_EN,
    HEADER_CATEGORY,
    HEADER_GERMAN_CATEGORY,
    HEADER_ACCOUNT,
    HEADER_PASSWORD,
    HEADER_DECLARED_WEIGHT,
    HEADER_FETCHED_WEIGHT,
]


@dataclass
class LoginTask:
    excel_path: Path
    sheet_name: str
    row_index: int
    authorized_representative: str
    weee_number: str
    company_name_cn: str
    company_name_en: str
    german_category: str
    account: str
    password: str


def log_info(message: str) -> None:
    print(f"[BUE1-EAR] {message}", flush=True)


def log_error(message: str) -> None:
    print(f"[BUE1-EAR] {message}", file=sys.stderr, flush=True)


def load_config() -> dict:
    config_path_value = os.environ.get("YISI_CONFIG_PATH", "").strip()
    if config_path_value:
        config_path = Path(config_path_value)
    else:
        config_path = Path(__file__).parent.parent.parent / "configs" / "BUE1_ear_declaration_data_fetcher.json"

    with open(config_path, "r", encoding="utf-8") as file:
        return json.load(file)


def normalize_header(value: object) -> str:
    text = str(value or "")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return text.strip()


def normalize_text_for_match(value: object) -> str:
    text = str(value or "")
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_text_casefold(value: object) -> str:
    return normalize_text_for_match(value).casefold()


def resolve_excel_file(config: dict) -> Path:
    excel_path = Path(str(config.get("excelFilePath") or config.get("excelFolderPath") or "").strip())
    if not str(excel_path):
        raise ValueError("申报数据 Excel 文件路径不能为空。")
    if not excel_path.exists():
        raise FileNotFoundError(f"申报数据 Excel 文件不存在: {excel_path}")
    if not excel_path.is_file():
        raise FileNotFoundError(f"申报数据 Excel 路径不是有效文件: {excel_path}")
    if excel_path.suffix.lower() not in SUPPORTED_EXCEL_SUFFIXES:
        raise ValueError(f"申报数据 Excel 文件格式不受支持，仅支持 .xlsx/.xlsm: {excel_path}")
    return excel_path


def ensure_excel_read_write_access(excel_path: Path) -> None:
    try:
        with open(excel_path, "rb") as file:
            file.read(1)
    except Exception as error:
        raise RuntimeError(f"当前 Excel 无法读取，请检查文件是否存在或是否有权限访问: {excel_path}") from error

    try:
        with open(excel_path, "r+b"):
            pass
    except Exception as error:
        raise RuntimeError(build_write_back_error_message(error, excel_path)) from error


def resolve_report_period(config: dict) -> tuple[str, str]:
    report_year = str(config.get("reportYear") or "").strip()
    report_month_german = str(config.get("reportMonthGerman") or "").strip()
    if not report_year:
        raise ValueError("配置中的年份不能为空。")
    if not report_month_german:
        raise ValueError("配置中的德语月份不能为空。")
    return report_year, report_month_german


def ensure_expected_headers(header_map: dict[str, int], excel_path: Path, sheet_name: str) -> None:
    missing_headers = [header for header in EXPECTED_HEADERS if header not in header_map]
    if missing_headers:
        raise ValueError(
            f"Excel 表头不完整: {excel_path.name} / {sheet_name}; 缺少: {', '.join(missing_headers)}"
        )


def load_excel_runtime(excel_path: Path, report_year: str, report_month_german: str):
    values_workbook = load_workbook(
        excel_path,
        read_only=True,
        data_only=True,
        keep_vba=excel_path.suffix.lower() == ".xlsm",
    )
    values_worksheet = values_workbook.active

    header_row = next(values_worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        values_workbook.close()
        raise ValueError(f"Excel 首行为空: {excel_path}")

    header_map = {
        normalize_header(value): index
        for index, value in enumerate(header_row)
        if normalize_header(value)
    }
    ensure_expected_headers(header_map, excel_path, values_worksheet.title)

    tasks: list[LoginTask] = []
    account_index = header_map[HEADER_ACCOUNT]
    password_index = header_map[HEADER_PASSWORD]
    german_category_index = header_map[HEADER_GERMAN_CATEGORY]
    representative_index = header_map[HEADER_AUTHORIZED_REP]
    weee_number_index = header_map[HEADER_WEEE_NUMBER]
    company_name_cn_index = header_map[HEADER_COMPANY_NAME_CN]
    company_name_en_index = header_map[HEADER_COMPANY_NAME_EN]
    fetched_weight_index = header_map[HEADER_FETCHED_WEIGHT]

    try:
        for row_index, row in enumerate(values_worksheet.iter_rows(min_row=2, values_only=True), start=2):
            account = normalize_text_for_match(row[account_index] if account_index < len(row) else "")
            password = normalize_text_for_match(row[password_index] if password_index < len(row) else "")
            fetched_weight = normalize_text_for_match(
                row[fetched_weight_index] if fetched_weight_index < len(row) else ""
            )
            if fetched_weight:
                log_info(f"跳过已有官网数据行: {excel_path.name}#{row_index}, 官网数据={fetched_weight}")
                continue
            if not account or not password:
                log_info(f"跳过空账号/密码行: {excel_path.name}#{row_index}")
                continue

            tasks.append(
                LoginTask(
                    excel_path=excel_path,
                    sheet_name=values_worksheet.title,
                    row_index=row_index,
                    authorized_representative=normalize_text_for_match(
                        row[representative_index] if representative_index < len(row) else ""
                    ),
                    weee_number=normalize_text_for_match(
                        row[weee_number_index] if weee_number_index < len(row) else ""
                    ),
                    company_name_cn=normalize_text_for_match(
                        row[company_name_cn_index] if company_name_cn_index < len(row) else ""
                    ),
                    company_name_en=normalize_text_for_match(
                        row[company_name_en_index] if company_name_en_index < len(row) else ""
                    ),
                    german_category=normalize_text_for_match(
                        row[german_category_index] if german_category_index < len(row) else ""
                    ),
                    account=account,
                    password=password,
                )
            )
    finally:
        values_workbook.close()

    runtime = {
        "tasks": tasks,
        "outputColumn": header_map[HEADER_FETCHED_WEIGHT] + 1,
        "reportYear": report_year,
        "reportMonthGerman": report_month_german,
    }
    return runtime


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)


def sanitize_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_") or "row"


def launch_isolated_browser_context(playwright, browser_profile_dir: Path) -> tuple[BrowserContext, str]:
    browser_profile_dir.mkdir(parents=True, exist_ok=True)
    common_args = [
        "--start-maximized",
        "--new-window",
        "--no-first-run",
        "--no-default-browser-check",
    ]

    try:
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=str(browser_profile_dir),
            channel="chrome",
            headless=True,
            viewport={"width": 1440, "height": 960},
            args=common_args,
        )
        return context, "chrome"
    except Exception as error:
        log_info(f"独立 Chrome 实例启动失败，改用 Playwright Chromium: {error}")
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=str(browser_profile_dir),
            headless=True,
            viewport={"width": 1440, "height": 960},
            args=common_args,
        )
        return context, "chromium"


def get_or_create_page(context: BrowserContext) -> Page:
    return context.pages[0] if context.pages else context.new_page()


def get_login_frame(page: Page):
    page.goto(EAR_PORTAL_URL, wait_until="domcontentloaded")
    page.wait_for_load_state("networkidle", timeout=20000)

    for frame in page.frames:
        if frame.url and LOGIN_FRAME_PATTERN.search(frame.url):
            return frame

    frame_locator = page.locator('iframe[src*="portalaufgaben.jsf"], frame[src*="portalaufgaben.jsf"]').first
    frame_locator.wait_for(state="attached", timeout=20000)
    frame_handle = frame_locator.element_handle()
    if frame_handle:
        frame = frame_handle.content_frame()
        if frame:
            return frame

    raise RuntimeError("未找到 EAR 登录 frame: /ear-portal/secure/common/portalaufgaben.jsf")


def get_portal_frame(page: Page):
    for _ in range(10):
        for frame in page.frames:
            if frame.url and LOGIN_FRAME_PATTERN.search(frame.url):
                return frame

        frame_locator = page.locator('iframe[src*="portalaufgaben.jsf"], frame[src*="portalaufgaben.jsf"]').first
        try:
            frame_locator.wait_for(state="attached", timeout=2000)
        except PlaywrightTimeoutError:
            page.wait_for_timeout(1000)
            continue

        frame_handle = frame_locator.element_handle()
        if frame_handle:
            frame = frame_handle.content_frame()
            if frame:
                return frame

        page.wait_for_timeout(1000)

    raise RuntimeError("登录后未找到 EAR portal frame: /ear-portal/secure/common/portalaufgaben.jsf")


def is_locator_visible(locator: Locator, timeout_ms: int = 5000) -> bool:
    try:
        locator.wait_for(state="visible", timeout=timeout_ms)
        return True
    except PlaywrightTimeoutError:
        return False


def get_ist_input_link(container) -> Locator:
    return container.locator(
        'a[href="/ear-portal/secure/hersteller/ist-inputmeldung/ist-inputmeldungen.jsf"], '
        'a[href*="/secure/hersteller/ist-inputmeldung/ist-inputmeldungen.jsf"], '
        'a:has-text("Ist-Inputmitteilungen")'
    ).first


def get_company_search_input(container) -> Locator:
    return container.locator('input[name="vertretener"].herstellerTextInput').first


def get_company_search_button(container) -> Locator:
    return container.locator(
        'button.btn[value="➡︎"], button.btn[value="➡"], button.btn:has-text("➡︎"), button.btn:has-text("➡")'
    ).first


def get_company_exit_button(container) -> Locator:
    return container.locator('span:has-text("✖︎"), span:has-text("✖")').first


def is_ist_input_page_ready(container) -> bool:
    return (
        is_locator_visible(container.locator("text=Meldejahr").first, timeout_ms=2000)
        or is_locator_visible(container.locator("text=Meldegewicht").first, timeout_ms=2000)
    )


def company_search_required(container) -> bool:
    return is_locator_visible(get_company_search_input(container), timeout_ms=3000)


def wait_for_ist_input_or_company_search(
    page: Page,
    attempts: int = 20,
    interval_ms: int = 150,
    include_company_search: bool = False,
):
    for _ in range(attempts):
        try:
            portal_frame = get_portal_frame(page)
        except Exception:
            page.wait_for_timeout(interval_ms)
            continue

        if is_ist_input_page_ready(portal_frame):
            return portal_frame
        if is_locator_visible(get_ist_input_link(portal_frame), timeout_ms=200):
            return portal_frame
        if include_company_search and is_locator_visible(get_company_search_input(portal_frame), timeout_ms=200):
            return portal_frame

        page.wait_for_timeout(interval_ms)

    return None


def exit_current_company(page: Page, portal_frame, log_prefix: str = "") -> bool:
    exit_button = get_company_exit_button(portal_frame)
    if not is_locator_visible(exit_button, timeout_ms=5000):
        return False

    if log_prefix:
        log_info(f"{log_prefix} 找到当前公司退出按钮，准备退出当前公司")

    try:
        exit_button.click(timeout=10000)
    except Exception:
        try:
            exit_button.click(timeout=10000, force=True)
        except Exception:
            return False

    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass
    page.wait_for_timeout(1500)

    refreshed_frame = get_portal_frame(page)
    if company_search_required(refreshed_frame):
        if log_prefix:
            log_info(f"{log_prefix} 已退出当前公司，返回公司搜索表单")
        return True
    return False


def enter_company_by_weee(page: Page, portal_frame, weee_number: str, log_prefix: str = "") -> bool:
    search_input = get_company_search_input(portal_frame)
    search_button = get_company_search_button(portal_frame)

    if not is_locator_visible(search_input, timeout_ms=5000):
        return False

    search_input.fill("")
    search_input.fill(weee_number)
    if log_prefix:
        log_info(f"{log_prefix} 检测到需要搜索公司，已填入 WEEE号={weee_number}")

    if not is_locator_visible(search_button, timeout_ms=5000):
        return False

    try:
        search_button.click(timeout=10000)
    except Exception:
        try:
            search_button.click(timeout=10000, force=True)
        except Exception:
            return False

    if log_prefix:
        log_info(f"{log_prefix} 已点击公司搜索进入按钮")

    refreshed_frame = wait_for_ist_input_or_company_search(
        page,
        attempts=25,
        interval_ms=120,
        include_company_search=False,
    )
    if refreshed_frame is not None:
        return True

    return True


def open_ist_input_page(page: Page, portal_frame, log_prefix: str = "") -> bool:
    if IST_INPUT_PATH_FRAGMENT in page.url or is_ist_input_page_ready(portal_frame):
        if log_prefix:
            log_info(f"{log_prefix} 已位于 Ist-Inputmitteilungen 页面")
        return True

    ist_link = get_ist_input_link(portal_frame)
    for _ in range(3):
        if not is_locator_visible(ist_link, timeout_ms=1500):
            page.wait_for_timeout(300)
            continue

        try:
            ist_link.scroll_into_view_if_needed(timeout=5000)
        except Exception:
            pass

        if log_prefix:
            log_info(f"{log_prefix} 找到 Ist-Inputmitteilungen 按钮")

        try:
            ist_link.click(timeout=10000)
        except Exception:
            try:
                ist_link.click(timeout=10000, force=True)
            except Exception:
                page.wait_for_timeout(300)
                continue

        try:
            page.wait_for_load_state("networkidle", timeout=20000)
        except Exception:
            pass

        refreshed_frame = get_portal_frame(page)
        if is_ist_input_page_ready(refreshed_frame):
            if log_prefix:
                log_info(f"{log_prefix} 成功点击 Ist-Inputmitteilungen 按钮")
            return True
        page.wait_for_timeout(300)

    return False


def categories_match(row_category: str, target_category: str) -> bool:
    row_norm = normalize_text_casefold(row_category)
    target_norm = normalize_text_casefold(target_category)
    if not row_norm or not target_norm:
        return False
    return row_norm == target_norm or row_norm in target_norm or target_norm in row_norm


def extract_matching_weight(
    container,
    report_year: str,
    report_month_german: str,
    german_category: str,
    log_prefix: str = "",
) -> str:
    row_sets = [
        container.locator("table tbody tr"),
        container.locator("tbody tr"),
        container.locator("tr"),
    ]

    report_year_norm = normalize_text_casefold(report_year)
    report_month_norm = normalize_text_casefold(report_month_german)

    for row_locator in row_sets:
        row_count = row_locator.count()
        if row_count == 0:
            continue

        for index in range(row_count):
            row = row_locator.nth(index)
            cells = [normalize_text_for_match(text) for text in row.locator("td").all_inner_texts()]
            if len(cells) < 4:
                continue

            row_year = normalize_text_casefold(cells[0])
            row_month = normalize_text_casefold(cells[1])
            row_category = cells[2]
            row_weight = cells[3]

            if row_year != report_year_norm:
                continue
            if row_month != report_month_norm:
                continue
            if not categories_match(row_category, german_category):
                continue

            if log_prefix:
                log_info(
                    f"{log_prefix} 找到目标数据: 年份={cells[0]}, 月份={cells[1]}, "
                    f"德语类目={row_category}, 重量={row_weight}"
                )
            return row_weight

    if log_prefix:
        log_info(
            f"{log_prefix} 未找到目标数据: 年份={report_year}, 月份={report_month_german}, "
            f"德语类目={german_category}"
        )
    return ""


def attempt_logout(page: Page, portal_frame=None, log_prefix: str = "") -> tuple[bool, str]:
    logout_link = None
    if portal_frame is not None:
        logout_link = portal_frame.locator("#navbarFormId\\:idSignout").first
        if not is_locator_visible(logout_link, timeout_ms=3000):
            logout_link = None

    if logout_link is None:
        logout_link = page.locator("#navbarFormId\\:idSignout").first
        try:
            logout_link.wait_for(state="visible", timeout=5000)
        except PlaywrightTimeoutError:
            return False, "logout_button_not_found"

    try:
        if log_prefix:
            log_info(f"{log_prefix} 找到退出登录按钮，准备点击")
        logout_link.click()
        try:
            page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            pass
        page.wait_for_timeout(1500)
        if log_prefix:
            log_info(f"{log_prefix} 成功点击退出登录按钮")
        return True, ""
    except Exception as error:
        return False, str(error)


def attempt_login_and_fetch(
    page: Page,
    task: LoginTask,
    screenshot_dir: Path,
    report_year: str,
    report_month_german: str,
    reuse_session: bool = False,
    previous_route: str = "",
) -> dict:
    screenshot_dir.mkdir(parents=True, exist_ok=True)

    result = {
        "excel": str(task.excel_path),
        "sheet": task.sheet_name,
        "rowIndex": task.row_index,
        "reportYear": report_year,
        "reportMonthGerman": report_month_german,
        "germanCategory": task.german_category,
        "account": task.account,
        "status": "unknown",
        "route": "",
        "matchedWeight": "",
        "pageUrl": "",
        "loginFormStillVisible": None,
        "logoutAttempted": False,
        "logoutSucceeded": False,
        "logoutError": "",
        "screenshotPath": "",
        "error": "",
    }

    try:
        log_prefix = (
            f"第{task.row_index}行 账号={task.account or '-'} "
            f"月份={report_month_german or '-'} 德语类目={task.german_category or '-'}"
        )
        login_form_still_visible = False
        portal_frame = None

        if reuse_session:
            log_info(f"{log_prefix} 账号密码与上一行相同，复用当前登录会话")
        else:
            frame = get_login_frame(page)

            account_input = frame.locator("#inputUserId")
            password_input = frame.locator("#inputPassword")
            login_button = frame.locator("button[type='submit'].btn.btn-primary")

            account_input.wait_for(state="visible", timeout=20000)
            password_input.wait_for(state="visible", timeout=20000)
            login_button.wait_for(state="visible", timeout=20000)

            account_input.fill(task.account)
            password_input.fill(task.password)
            login_button.click()
            log_info(f"{log_prefix} 已提交登录")

            portal_frame = wait_for_ist_input_or_company_search(
                page,
                attempts=25,
                interval_ms=120,
                include_company_search=True,
            )

            try:
                login_form_still_visible = account_input.is_visible(timeout=1000)
            except Exception:
                login_form_still_visible = False

        result["pageUrl"] = page.url
        result["loginFormStillVisible"] = login_form_still_visible

        if login_form_still_visible:
            result["status"] = "login_form_still_visible"
            log_info(f"{log_prefix} 登录后仍停留在登录页")
        else:
            portal_frame = get_portal_frame(page)
            route = "direct_company"

            if reuse_session and previous_route == "searched_company":
                exited_company = exit_current_company(page, portal_frame, log_prefix=log_prefix)
                if not exited_company:
                    result["route"] = "company_selection_required"
                    result["status"] = "company_selection_required"
                    result["error"] = "复用同账号会话时，无法退出当前公司。"
                    log_info(f"{log_prefix} 复用搜索公司分支时，退出当前公司失败")
                    return result
                portal_frame = get_portal_frame(page)
                company_entered = enter_company_by_weee(page, portal_frame, task.weee_number, log_prefix=log_prefix)
                if not company_entered:
                    result["route"] = "company_selection_required"
                    result["status"] = "company_selection_required"
                    result["error"] = "需要搜索公司，但未能通过 WEEE号进入目标公司。"
                    log_info(f"{log_prefix} 退出当前公司后，未能通过 WEEE号进入目标公司")
                    return result
                portal_frame = get_portal_frame(page)
                route = "searched_company"
                log_info(f"{log_prefix} 已进入目标公司，开始点击 Ist-Inputmitteilungen")
                ist_input_opened = open_ist_input_page(page, portal_frame, log_prefix=log_prefix)
            elif reuse_session and previous_route == "direct_company" and is_ist_input_page_ready(portal_frame):
                ist_input_opened = True
                log_info(f"{log_prefix} 复用当前 Ist-Inputmitteilungen 页面，直接检索当前行数据")
            else:
                if company_search_required(portal_frame):
                    company_entered = enter_company_by_weee(page, portal_frame, task.weee_number, log_prefix=log_prefix)
                    if not company_entered:
                        result["route"] = "company_selection_required"
                        result["status"] = "company_selection_required"
                        result["error"] = "需要搜索公司，但未能通过 WEEE号进入目标公司。"
                        log_info(f"{log_prefix} 检测到需要搜索公司，但未能通过 WEEE号进入目标公司")
                        return result
                    portal_frame = get_portal_frame(page)
                    route = "searched_company"

                log_info(f"{log_prefix} 已进入 portal frame，开始寻找 Ist-Inputmitteilungen 按钮")
                ist_input_opened = open_ist_input_page(page, portal_frame, log_prefix=log_prefix)
            result["pageUrl"] = page.url

            if ist_input_opened:
                result["route"] = route
                portal_frame = get_portal_frame(page)
                log_info(f"{log_prefix} 开始寻找目标数据")
                matched_weight = extract_matching_weight(
                    portal_frame,
                    report_year,
                    report_month_german,
                    task.german_category,
                    log_prefix=log_prefix,
                )
                result["matchedWeight"] = matched_weight
                if matched_weight:
                    result["status"] = "weight_found"
                    result["error"] = ""
                else:
                    result["status"] = "matching_entry_not_found"
                    result["error"] = "未找到符合条件的数据"
            else:
                result["route"] = "company_selection_required"
                result["status"] = "company_selection_required"
                result["error"] = "检测到当前账号登录后未直接进入公司，需补充公司搜索与进入逻辑。"
                log_info(f"{log_prefix} 未能打开 Ist-Inputmitteilungen，当前按多公司账号处理")

        screenshot_name = sanitize_filename(f"{task.excel_path.stem}_{task.sheet_name}_row_{task.row_index}.png")
        screenshot_path = screenshot_dir / screenshot_name
        page.screenshot(path=str(screenshot_path), full_page=True)
        result["screenshotPath"] = str(screenshot_path)

        return result
    except PlaywrightTimeoutError as error:
        result["status"] = "timeout"
        result["error"] = str(error)
        return result
    except Exception as error:
        result["status"] = "failed"
        result["error"] = str(error)
        return result


def write_weight_back(row_index: int, output_column: int, weight_value: str, excel_path: Path) -> None:
    workbook = load_workbook(
        excel_path,
        read_only=False,
        data_only=False,
        keep_vba=excel_path.suffix.lower() == ".xlsm",
    )
    temp_path = excel_path.with_name(f"{excel_path.stem}.tmp.{uuid4().hex}{excel_path.suffix}")
    try:
        worksheet = workbook.active
        worksheet.cell(row=row_index, column=output_column, value=weight_value)
        workbook.save(temp_path)
    finally:
        workbook.close()

    try:
        os.replace(temp_path, excel_path)
    except Exception:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
        raise


def build_write_back_error_message(error: Exception, excel_path: Path) -> str:
    error_text = str(error)
    if isinstance(error, PermissionError) or "Permission denied" in error_text:
        return f"当前 Excel 无法写入，请检查是否正在被打开中: {excel_path}"
    return f"回填 Excel 失败，请检查文件是否可写: {excel_path}; {error_text}"


def build_manifest(
    config: dict,
    excel_files: list[Path],
    tasks: list[LoginTask],
    browser_profile_dir: Path,
    browser_engine: str,
    report_year: str,
    report_month_german: str,
) -> dict:
    runtime_dir = os.environ.get("YISI_RUNTIME_DIR", "").strip()
    run_id = os.environ.get("YISI_RUN_ID", "").strip()

    return {
        "generatedAt": datetime.now().isoformat(),
        "department": "BUE1",
        "tool": "ear_declaration_data_fetcher",
        "runId": run_id,
        "runtimeDir": runtime_dir,
        "excelFilePath": config.get("excelFilePath") or config.get("excelFolderPath", ""),
        "reportYear": report_year,
        "reportMonthGerman": report_month_german,
        "browserEngine": browser_engine,
        "browserProfileDir": str(browser_profile_dir),
        "expectedHeaders": EXPECTED_HEADERS,
        "excelFiles": [str(path) for path in excel_files],
        "taskCount": len(tasks),
        "notes": [
            "当前版本会启动独立的有头浏览器实例。",
            "浏览器使用独立 user data dir，不会复用你本地默认 Chrome 个人资料。",
            "每条记录会从 Excel 的“账号”“密码”列读取登录信息。",
            "年份和德语月份从配置项读取，不再从 Excel 的月份列读取。",
            "如果登录后未直接出现 Ist-Inputmitteilungen，则会标记为待补充公司搜索逻辑。",
        ],
    }


def main() -> int:
    try:
        log_info("开始执行 EAR 官网申报数据抓取流程。")
        config = load_config()

        excel_file_path = resolve_excel_file(config)
        ensure_excel_read_write_access(excel_file_path)
        log_info(f"Excel 读写权限检查通过: {excel_file_path}")
        report_year, report_month_german = resolve_report_period(config)
        runtime = load_excel_runtime(excel_file_path, report_year, report_month_german)
        tasks: list[LoginTask] = runtime["tasks"]
        output_column = runtime["outputColumn"]

        if not tasks:
            raise ValueError("没有可执行的登录任务，请检查 Excel 中的账号和密码列是否为空。")

        artifact_dir = Path(os.environ.get("YISI_ARTIFACT_DIR", "") or excel_file_path.parent)
        screenshot_dir = Path(os.environ.get("YISI_SCREENSHOT_DIR", "") or (artifact_dir / "screenshots"))
        browser_profile_dir = Path(
            os.environ.get("YISI_BROWSER_PROFILE_DIR", "") or (artifact_dir / "browser-profile")
        )
        manifest_path = artifact_dir / "ear_login_manifest.json"
        report_path = artifact_dir / "ear_login_report.json"

        results: list[dict] = []
        with sync_playwright() as playwright:
            context, browser_engine = launch_isolated_browser_context(playwright, browser_profile_dir)
            active_page = None
            active_credentials: tuple[str, str] | None = None
            active_session_reusable = False
            active_session_route = ""
            try:
                page = get_or_create_page(context)
                page.goto("about:blank")

                manifest = build_manifest(
                    config,
                    [excel_file_path],
                    tasks,
                    browser_profile_dir,
                    browser_engine,
                    report_year,
                    report_month_german,
                )
                write_json(manifest_path, manifest)

                log_info(f"已启动独立 {browser_engine} 实例，专用 profile: {browser_profile_dir}")
                log_info(f"共读取到 {len(tasks)} 条登录任务。")

                for task in tasks:
                    task_log_prefix = (
                        f"第{task.row_index}行 账号={task.account or '-'} "
                        f"月份={report_month_german or '-'} 德语类目={task.german_category or '-'}"
                    )
                    current_credentials = (task.account, task.password)
                    reuse_session = (
                        active_page is not None
                        and active_credentials == current_credentials
                        and active_session_reusable
                    )

                    if not reuse_session:
                        if active_page is not None:
                            log_info(f"{task_log_prefix} 检测到账号切换，准备退出上一账号")
                            previous_portal_frame = None
                            try:
                                previous_portal_frame = get_portal_frame(active_page)
                            except Exception:
                                previous_portal_frame = None
                            logout_succeeded, logout_error = attempt_logout(
                                active_page,
                                portal_frame=previous_portal_frame,
                                log_prefix=f"账号={active_credentials[0] if active_credentials else '-'}",
                            )
                            if not logout_succeeded:
                                log_info(f"{task_log_prefix} 上一账号退出失败: {logout_error}")
                            active_page.close()

                        active_page = context.new_page()
                        active_credentials = current_credentials
                        active_session_reusable = False
                        active_session_route = ""

                    log_info(
                        f"账号={task.account or '-'}, 密码={task.password or '-'}, "
                        f"月份={report_month_german or '-'}, 德语类目={task.german_category or '-'}"
                    )
                    result = attempt_login_and_fetch(
                        active_page,
                        task,
                        screenshot_dir,
                        report_year,
                        report_month_german,
                        reuse_session=reuse_session,
                        previous_route=active_session_route,
                    )
                    active_session_reusable = result["status"] in {"weight_found", "matching_entry_not_found"}
                    active_session_route = result.get("route", "")

                    write_back_value = result.get("matchedWeight")
                    if not write_back_value and result.get("status") == "matching_entry_not_found":
                        write_back_value = MISSING_MATCH_WRITEBACK_VALUE

                    if write_back_value:
                        try:
                            write_weight_back(
                                task.row_index,
                                output_column,
                                write_back_value,
                                excel_file_path,
                            )
                            result["writeBackSucceeded"] = True
                            result["writeBackError"] = ""
                            log_info(
                                f"{task_log_prefix} 已将官网数据 {write_back_value} 回填到 Excel 第{task.row_index}行"
                            )
                        except Exception as error:
                            friendly_error = build_write_back_error_message(error, excel_file_path)
                            result["writeBackSucceeded"] = False
                            result["writeBackError"] = friendly_error
                            log_info(f"{task_log_prefix} 回填 Excel 失败: {friendly_error}")
                            if result["status"] == "weight_found":
                                result["status"] = "write_back_failed"
                            results.append(result)
                            raise RuntimeError(friendly_error) from error
                    else:
                        result["writeBackSucceeded"] = False
                        result["writeBackError"] = ""
                        log_info(f"{task_log_prefix} 未获取到官网重量，未执行回填")

                    results.append(result)
            finally:
                if active_page is not None:
                    final_portal_frame = None
                    try:
                        final_portal_frame = get_portal_frame(active_page)
                    except Exception:
                        final_portal_frame = None
                    logout_succeeded, logout_error = attempt_logout(
                        active_page,
                        portal_frame=final_portal_frame,
                        log_prefix=f"账号={active_credentials[0] if active_credentials else '-'}",
                    )
                    if not logout_succeeded:
                        log_info(f"账号={active_credentials[0] if active_credentials else '-'} 最终退出失败: {logout_error}")
                    active_page.close()
                context.close()

        report_payload = {
            "generatedAt": datetime.now().isoformat(),
            "total": len(results),
            "results": results,
        }
        write_json(report_path, report_payload)

        success_count = sum(1 for item in results if item["status"] == "weight_found")
        log_info(f"抓取流程执行完成，结果文件: {report_path}")
        print("EAR 抓取流程执行完成")
        print(f"登录任务总数: {len(results)}")
        print(f"成功抓取重量数: {success_count}")
        print(f"清单文件: {manifest_path}")
        print(f"结果文件: {report_path}")
        return 0
    except Exception as error:
        message = f"执行失败: {error}"
        log_info(message)
        log_error(message)
        traceback.print_exc(file=sys.stdout)
        return 1


if __name__ == "__main__":
    sys.exit(main())
